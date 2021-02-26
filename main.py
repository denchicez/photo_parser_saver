import urllib.request
import requests
from bs4 import BeautifulSoup
import xlrd
from openpyxl.workbook import Workbook
from openpyxl import load_workbook
def open_xls_as_xlsx(filename):
    book = xlrd.open_workbook(filename)
    index = 0
    nrows, ncols = 0, 0
    while nrows * ncols == 0:
        sheet = book.sheet_by_index(index)
        nrows = sheet.nrows
        ncols = sheet.ncols
        index += 1
    book1 = Workbook()
    sheet1 = book1.get_active_sheet()
    for row in range(1, nrows):
        for col in range(1, ncols):
            sheet1.cell(row=row, column=col).value = sheet.cell_value(row, col)
    return book1
errors = set()
def get_html(url, params=None):
    r = requests.get(url, params=params)
    return r.text 
def save_images(URL,name,PATH):
    try:
        catalog = ''
        counter = 0
        for symbol in URL:
            if(symbol=='/'):
                counter+=1
            if(counter==3):
                break
            catalog = catalog + symbol
        html = get_html(URL)
        soup = BeautifulSoup(html,'lxml')
        if(URL.find('fkniga')!=-1):
            soup = soup.find('div',class_='section section--goodCard')
            soup = soup.find('div',class_='swiper-wrapper')
            soup = soup.find('div')
        if(URL.find('sunnydress')!=-1):
            soup = soup.find('div',class_='multizoom1 thumbs')
            soup = soup.find('a')
        images_soup = soup.find_all('img')
        images_src_set = set()
        suffix_set = set()
        for image_soup in images_soup:
            try:
                src = image_soup.get('src')
                if(src.find('http')!=-1):
                    continue
                suffix = ''
                for i in range(len(src)-1,0,-1):
                    if(src[i]!='/'):
                        suffix = src[i]+suffix
                    else:
                        break
                if((suffix in suffix_set) == False):
                    suffix_set.add(suffix)
                    images_src_set.add(src)
            except:
                print('Что-то с ссылочкой криво')
        if(len(images_src_set)==0):
            errors.add(URL)
        else:
            image_href = catalog+list(images_src_set)[-1]
            if(image_href.find('bG9jYWw6Ly8vbWVkaWEvbm8taW1hZ2UtcG5nLTItb3JpZ2luYWwucG5n.jpg')!=-1):
                errors.add(URL)
            else:
                try:
                    resource = urllib.request.urlopen(image_href)
                    out = open(f"{PATH}/{name}.jpg", 'wb')
                    out.write(resource.read())
                    out.close()
                except Exception as e:
                    print(URL)
                    print(e)
                    errors.add(URL)
    except Exception as e:
        print(URL)
        print(e)
        errors.add(URL)
urls = list()
names = list()
if __name__ == '__main__':
    print('Введите путь до папки в которую нужно вывести данные')
    PATH = input() # C:\Users\evroz\Desktop\berezok\прайс1
    print('Введите путь до файла от куда брать данные')
    PATH_EXCEL = input() # C:\Users\evroz\Desktop\berezok\Прайс_образец_1.xls
    if(PATH_EXCEL.find('xlsx')==-1):
        try:
            excel_df = xlrd.open_workbook(PATH_EXCEL)
            sheet = excel_df.sheet_by_index(0)
            hyperlinks = sheet.hyperlink_list
            for i in range(0,len(hyperlinks)):
                if((urls.count(hyperlinks[i].url_or_path))==0):
                    urls.append(hyperlinks[i].url_or_path)
                    new_name = hyperlinks[i].desc
                    new_name = new_name.replace('  ',' ')
                    names.append(new_name)
        except:
            excel_df = load_workbook(PATH_EXCEL)
            sheet = excel_df.worksheets[0]
            column = sheet['A']
            for cell in column:
                cellname = cell.value
                cellhref = cell.hyperlink.target
                if(cellhref==None):
                    errors.add(cellname)
                else:
                    if((urls.count(cellhref))==0):
                        urls.append(cellhref)
                        new_name = cellname
                        new_name = new_name.replace('  ',' ')
                        names.append(cellname)
    else:
        try:
            excel_df = load_workbook(PATH_EXCEL)
            sheet = excel_df.worksheets[0]
            column = sheet['A']
            for cell in column:
                cellname = cell.value
                if(cellhref==None):
                    errors.add(cellname)
                else:
                    cellhref = cell.hyperlink.target
                    if((urls.count(cellhref))==0):
                        urls.append(cellhref)
                        new_name = cellname
                        new_name = new_name.replace('  ',' ')
                        names.append(cellname)
        except:
            excel_df = xlrd.open_workbook(PATH_EXCEL)
            sheet = excel_df.sheet_by_index(0)
            hyperlinks = sheet.hyperlink_list
            for i in range(0,len(hyperlinks)):
                if((urls.count(hyperlinks[i].url_or_path))==0):
                    urls.append(hyperlinks[i].url_or_path)
                    new_name = hyperlinks[i].desc
                    new_name = new_name.replace('  ',' ')
                    names.append(new_name)
    for index in range(0,len(urls)):
        save_images(list(urls)[index],list(names)[index],PATH)
    with open(f'{PATH}/errors.txt', 'w') as f:
        big_error = ''
        for i in errors:
            big_error=big_error+str(i)+'\n'
        f.write(big_error)